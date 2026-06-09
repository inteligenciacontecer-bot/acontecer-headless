#!/usr/bin/env python3
"""
scraper_opendata.py
Sincroniza SALARIOS y COMBUSTIBLE de los diputados desde el NUEVO portal de
datos abiertos de la Asamblea (migrado a /pa/datosabiertos/).

- Salarios:   biblioteca "Documentos", archivo "*Salario Diputados*.xlsx" mas reciente.
              Salario bruto = Dieta + Gastos de Representacion + Reintegros.
- Combustible: lista "transporte_vehiculos", "*Uso_Transporte.xlsx" mas reciente.
              Suma de la columna COMBUSTIBLE por diputado en el periodo.

Toma SIEMPRE el archivo mas reciente -> cuando la Asamblea publique datos del
periodo 2026-2030, se cargan solos. Hace match a los diputados ACTIVOS por dos
apellidos (los archivos viejos no matchean -> no ensucian la base).

Cron sugerido: 30 7 * * 1 (semanal).
"""
import urllib.request, json, ssl, io, urllib.parse, re, unicodedata
from datetime import datetime, date

CTX = ssl._create_unverified_context()
OD = "https://www.asamblea.go.cr/pa/datosabiertos"
DB_PREFIX = "hake_"


def log(m):
    print(f"[OPENDATA {datetime.now().strftime('%H:%M:%S')}] {m}", flush=True)


# ── matching por 2 apellidos ──────────────────────────────────────────────────
def _norm(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").upper()
    s = s.replace("-", " ").replace(".", " ").replace(",", " ")
    return re.sub(r"\s+", " ", s).strip()

_PART = {"DE", "LA", "LOS", "LAS", "DEL", "Y", "G", "DIP", "DIPUTADO", "DIPUTADA"}

def _apellidos(nombre):
    toks = [t for t in _norm(nombre).split() if t not in _PART and len(t) > 1]
    return (toks[-2], toks[-1]) if len(toks) >= 2 else None

def match(nombre, diputados):
    nm = _norm(re.sub(r"\(.*?\)", "", nombre))  # quitar (NR), (PLN)...
    cands = []
    for did, nomb, a1, a2 in diputados:
        if re.search(rf"\b{re.escape(a1)}\b", nm) and re.search(rf"\b{re.escape(a2)}\b", nm):
            cands.append(did)
    return cands[0] if len(cands) == 1 else (cands[0] if cands else None)


# ── SharePoint helpers ────────────────────────────────────────────────────────
def fj(url):
    req = urllib.request.Request(url, headers={"Accept": "application/json;odata=verbose", "User-Agent": "Mozilla/5.0"})
    return json.loads(urllib.request.urlopen(req, timeout=40, context=CTX).read())

def descargar(rel):
    url = "https://www.asamblea.go.cr" + "/".join(urllib.parse.quote(p) for p in rel.split("/"))
    return urllib.request.urlopen(urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=90, context=CTX).read()

def archivos(lista, filtro):
    d = fj(f"{OD}/_api/web/lists/getbytitle('{urllib.parse.quote(lista)}')/items?$top=500&$select=FileLeafRef&$expand=File")
    out = []
    for it in d.get("d", {}).get("results", []):
        f = it.get("File") or {}
        nm = f.get("Name") or ""
        if filtro(nm):
            out.append((nm, f.get("ServerRelativeUrl"), f.get("TimeLastModified", "")))
    return out


def get_db():
    import pymysql
    with open("/var/www/cms.acontecer.co.cr/wp-config.php") as f:
        txt = f.read()
    def g(k):
        m = re.search(rf"define\s*\(\s*['\"]DB_{k}['\"]\s*,\s*['\"]([^'\"]+)['\"]\s*\)", txt)
        return m.group(1) if m else ""
    host, port = g("HOST") or "localhost", 3306
    if ":" in host:
        host, port = host.rsplit(":", 1); port = int(port)
    return pymysql.connect(host=host, port=port, user=g("USER"), password=g("PASSWORD"),
                           database=g("NAME"), charset="utf8mb4",
                           cursorclass=pymysql.cursors.DictCursor, autocommit=False)


def cargar_diputados(cur):
    cur.execute(f"SELECT id, nombre_completo FROM {DB_PREFIX}asamblea_diputados WHERE activo=1")
    out = []
    for d in cur.fetchall():
        ap = _apellidos(d["nombre_completo"])
        if ap:
            out.append((d["id"], d["nombre_completo"], ap[0], ap[1]))
    return out


def num(v):
    try:
        return float(v)
    except Exception:
        return 0.0


# ── SALARIOS ──────────────────────────────────────────────────────────────────
def sync_salarios(cur, diputados):
    files = archivos("Documentos", lambda n: "salario diputados" in n.lower() and n.lower().endswith("xlsx"))
    if not files:
        log("salarios: sin archivos"); return 0
    files.sort(key=lambda x: x[2], reverse=True)
    nm, rel, mod = files[0]
    log(f"salarios: archivo mas reciente = {nm} ({mod[:10]})")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(descargar(rel)), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    hdr = [str(c or "").strip().lower() for c in rows[0]]
    def col(*keys):
        for i, h in enumerate(hdr):
            if any(k in h for k in keys): return i
        return None
    ci = {"nombre": col("nombre"), "mes": col("mes"), "anio": col("año", "ano", "anio"),
          "dieta": col("dieta"), "gastos": col("gastos"), "rein": col("reintegro"), "bruto": col("bruto")}
    n = 0
    for r in rows[1:]:
        if ci["nombre"] is None or not r[ci["nombre"]]:
            continue
        did = match(str(r[ci["nombre"]]), diputados)
        if not did:
            continue
        bruto = num(r[ci["bruto"]]) if ci["bruto"] is not None else 0
        if bruto <= 0:
            bruto = num(r[ci["dieta"]]) + num(r[ci["gastos"]]) + (num(r[ci["rein"]]) if ci["rein"] is not None else 0)
        try:
            mes = int(r[ci["mes"]]); anio = int(r[ci["anio"]])
            periodo = date(anio, mes, 1)
        except Exception:
            continue
        cur.execute(
            f"""INSERT INTO {DB_PREFIX}asamblea_salarios (nombre, mes, salario_bruto)
                VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE salario_bruto=VALUES(salario_bruto)""",
            (str(r[ci["nombre"]])[:200], periodo, round(bruto, 2)))
        n += 1
    log(f"salarios: {n} filas de diputados ACTUALES")
    return n


# ── COMBUSTIBLE ───────────────────────────────────────────────────────────────
def sync_combustible(cur, diputados):
    files = archivos("transporte_vehiculos", lambda n: "transporte" in n.lower() and n.lower().endswith("xlsx"))
    if not files:
        log("combustible: sin archivos"); return 0
    files.sort(key=lambda x: x[2], reverse=True)
    nm, rel, mod = files[0]
    log(f"combustible: archivo mas reciente = {nm} ({mod[:10]})")
    mper = re.search(r"(20\d\d)[-_](\d{2})", nm)
    periodo = date(int(mper.group(1)), int(mper.group(2)), 1) if mper else date(int(mod[:4]), int(mod[5:7]), 1)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(descargar(rel)), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c or "").strip().lower() for c in rows[0]]
    def col(*keys):
        for i, h in enumerate(hdr):
            if any(k in h for k in keys): return i
        return None
    c_dip, c_comb = col("diputado"), col("combustible")
    if c_dip is None or c_comb is None:
        log("combustible: columnas no encontradas"); return 0
    agg = {}
    for r in rows[1:]:
        if not r[c_dip]:
            continue
        did = match(str(r[c_dip]), diputados)
        if not did:
            continue
        agg[did] = agg.get(did, 0) + num(r[c_comb])
    n = 0
    for did, monto in agg.items():
        cur.execute(f"SELECT nombre_completo FROM {DB_PREFIX}asamblea_diputados WHERE id=%s", (did,))
        nombre = cur.fetchone()["nombre_completo"]
        cur.execute(
            f"""INSERT INTO {DB_PREFIX}asamblea_gasolina (nombre, periodo, monto)
                VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE monto=VALUES(monto)""",
            (nombre[:200], periodo, round(monto, 2)))
        n += 1
    log(f"combustible: {n} diputados ACTUALES en {periodo}")
    return n


def main():
    db = get_db()
    with db.cursor() as cur:
        diputados = cargar_diputados(cur)
        log(f"Diputados activos: {len(diputados)}")
        try:
            sync_salarios(cur, diputados)
        except Exception as e:
            log(f"ERROR salarios: {e}")
        try:
            sync_combustible(cur, diputados)
        except Exception as e:
            log(f"ERROR combustible: {e}")
        db.commit()
    db.close()
    log("Finalizado (captura automatica cuando la Asamblea publique datos 2026+).")


if __name__ == "__main__":
    main()
